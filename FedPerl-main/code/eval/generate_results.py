import glob
from glob import glob

import numpy as np
import openpyxl
import torch
import torch.nn as nn
# pytorch libraries
from efficientnet_pytorch import EfficientNet
# sklearn libraries
from sklearn.metrics import precision_recall_fscore_support

from data.data_FL import SkinData

# Compute precision, recall, F-measure and support for each class.
# The precision is intuitively the ability of the classifier not to label as positive a sample that is negative.
# The recall is intuitively the ability of the classifier to find all the positive samples.
# The F-beta score can be interpreted as a weighted harmonic mean of the precision and recall.
# The F-beta score weights recall more than precision by a factor of beta.
# beta == 1.0 means recall and precision are equally important.
# The support is the number of occurrences of each class in y_true.

lesions_names = ['MEL', 'NV', 'BCC', 'AK', 'BKL', 'DF', 'VASC', 'SCC']
lesions_indexs = [0, 1, 2, 3, 4, 5, 6, 7]

num_classes = 8
batchSize = 16
data_path = '.......'
name = 'FedPerl'


def predict(model, loader):
    model.eval()
    y_gt = []
    y_preds = []
    with torch.no_grad():
        for batch_idx, sample_batched in enumerate(loader):
            X = sample_batched[0].type(torch.cuda.FloatTensor)
            y = sample_batched[1].type(torch.cuda.LongTensor)
            y_gt.append(y.cpu())
            y_pred = model(X)
            y_pred = np.argmax(y_pred.cpu(), axis=1)
            y_preds.append(y_pred)
    return np.concatenate(y_gt), np.concatenate(y_preds)


def calculate_scores_manually(y, predictions):
    """
    calculate scores

    Parameters:
       y: ground truths
       predictions: predictions

    Returns:
        None
    """
    TP = 0
    FP = 0
    TN = 0
    FN = 0

    for i in range(len(predictions)):
        if y[i] == predictions[i]:
            TP += 1
        if y[i] != predictions[i]:
            FP += 1
        if y[i] == predictions[i] == 0:
            TN += 1
        if predictions[i] == 0 and y[i] != predictions[i]:
            FN += 1

    # Sensitivity, hit rate, recall, or true positive rate
    TPR = TP / (TP + FN)
    # Specificity or true negative rate
    TNR = TN / (TN + FP)
    # Precision or positive predictive value
    PPV = TP / (TP + FP)
    # Negative predictive value
    NPV = TN / (TN + FN)
    # Fall out or false positive rate
    FPR = FP / (FP + TN)
    # False negative rate
    FNR = FN / (TP + FN)
    # False discovery rate
    FDR = FP / (TP + FP)
    # Overall accuracy
    ACC = (TP + TN) / (TP + FP + FN + TN)

    # F1 = 2 * (precision * recall) / (precision + recall)
    F1 = 2 * (PPV * TPR) / (PPV + TPR)


def calculate_scores(y, predictions):
    """
    calculate scores using skitlearn

    Parameters:
        y: ground truths
        predictions: predictions

    Returns:
        scores
    """
    acc = np.mean(np.equal(predictions, y))
    prf3 = precision_recall_fscore_support(y, predictions, average='weighted')

    return acc, prf3[0], prf3[1], prf3[2]


def generate_summary():
    """
    generate model summary and save it an excel sheet

    Parameters:
        None

    Returns:
        None
    """
    models = []
    for model in glob.glob(model_path + '*.pt'):
        models.append(model)

    model = EfficientNet.from_pretrained('efficientnet-b0', num_classes=num_classes)
    num_ftrs = model._fc.in_features
    model._fc = nn.Linear(num_ftrs, num_classes)
    device = torch.device('cuda:0')
    model = nn.DataParallel(model)
    model = model.cuda()
    model = model.to(device)

    wb = openpyxl.load_workbook('All_Scores.xlsx')
    i = 1
    shift = 1
    for name in models:
        test_ds, val_ds = get_dataset(name)

        valid_loader = torch.utils.data.DataLoader(val_ds, batch_size=batchSize, shuffle='False', num_workers=4,
                                                   pin_memory=True)
        test_loader = torch.utils.data.DataLoader(test_ds, batch_size=batchSize, shuffle='False', num_workers=4,
                                                  pin_memory=True)

        ws = wb.get_sheet_by_name('Val')

        model.load_state_dict(torch.load(name))
        # for i in range(5):
        model.eval()

        y, predictions = predict(model, valid_loader)
        Acc, wPr, wR, wF = calculate_scores(y, predictions)

        cl = 'A' + str(i + shift)
        ws[cl] = name
        cl = 'B' + str(i + shift)
        ws[cl] = Acc
        cl = 'C' + str(i + shift)
        ws[cl] = wF
        cl = 'D' + str(i + shift)
        ws[cl] = wPr
        cl = 'E' + str(i + shift)
        ws[cl] = wR
        cl = 'F' + str(i + shift)

        ws = wb.get_sheet_by_name('Test')
        y, predictions = predict(model, test_loader)
        Acc, wPr, wR, wF = calculate_scores(y, predictions)

        cl = 'A' + str(i + shift)
        ws[cl] = name
        cl = 'B' + str(i + shift)
        ws[cl] = Acc
        cl = 'C' + str(i + shift)
        ws[cl] = wF
        cl = 'D' + str(i + shift)
        ws[cl] = wPr
        cl = 'E' + str(i + shift)
        ws[cl] = wR
        cl = 'F' + str(i + shift)
        i += 1

    wb.save('All_Scores.xlsx')


def get_dataset(name):
    cid = -1
    if 'Client0' in name:
        cid = 0
    elif 'Client1' in name:
        cid = 1
    elif 'Client2' in name:
        cid = 2
    elif 'Client3' in name:
        cid = 3
    elif 'Client4' in name:
        cid = 4
    elif 'Client5' in name:
        cid = 5
    elif 'Client6' in name:
        cid = 6
    elif 'Client7' in name:
        cid = 7
    elif 'Client8' in name:
        cid = 8
    elif 'Client9' in name:
        cid = 9

    data = SkinData(data_path, None)
    test_ds, val_ds = data.load_clients_test_val(cid)

    return test_ds, val_ds


if __name__ == '__main__':
    generate_summary()
